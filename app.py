# app.py
from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
import os

app = Flask(__name__)

MENU_XLSX = Path("menu.xlsx")
BILLS_XLSX = Path("bills.xlsx")
TABLES = ["Outside 1", "Outside 2", "Outside 3", "Inside 1", "Inside 2", "Inside 3", "Last 1", "Last 2"]

# -----------------------------
# Excel utilities (robust)
# -----------------------------
def ensure_menu_file():
    if not MENU_XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Item Name", "Price"])
        wb.save(MENU_XLSX)

def ensure_bills_file():
    if not BILLS_XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Bill No", "Date & Time", "Item Name", "Qty", "Rate", "Amount", "Total", "Payment Method", "Table"])
        wb.save(BILLS_XLSX)

def load_menu_from_xlsx():
    """
    Reads menu.xlsx and returns dict: { category: [ {name, price}, ... ] }
    If file not present, create with header and return {} (no sample hardcoded)
    """
    ensure_menu_file()
    wb = load_workbook(MENU_XLSX)
    sheet = wb.active
    data = {}
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            # header
            continue
        # take only first 3 columns to avoid unpack errors
        cells = list(row)[:3]
        if len(cells) < 2:
            continue
        cat = cells[0] if cells[0] is not None else ""
        name = cells[1] if cells[1] is not None else ""
        price = cells[2] if len(cells) >= 3 and cells[2] not in (None, "") else 0
        cat = str(cat).strip()
        name = str(name).strip()
        try:
            price = float(price)
        except:
            # ignore invalid price rows
            continue
        if not name:
            continue
        data.setdefault(cat or "Uncategorized", []).append({"name": name, "price": price})
    return data

def write_menu_to_xlsx(menu_dict):
    """
    Overwrite menu.xlsx with the provided menu_dict
    format: {category: [{name,price}, ...]}
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Category", "Item Name", "Price"])
    for cat, items in menu_dict.items():
        for it in items:
            ws.append([cat, it["name"], it["price"]])
    wb.save(MENU_XLSX)

def next_bill_no():
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX, read_only=True)
    sheet = wb.active
    max_bn = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            val = int(row[0])
            if val > max_bn:
                max_bn = val
        except:
            continue
    wb.close()
    return max_bn + 1

def append_bill_to_xlsx(bill_obj):
    """
    bill_obj structure:
      { billNo, dateTime(str), table, payment, total(float), items: [{name,qty,rate,amount}, ...] }
    Appends one row per item in bills.xlsx
    """
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX)
    sheet = wb.active
    bn = bill_obj.get("billNo", next_bill_no())
    dt = bill_obj.get("dateTime", datetime.now().isoformat())
    payment = bill_obj.get("payment", "")
    total = bill_obj.get("total", 0.0)
    table = bill_obj.get("table", "")
    for it in bill_obj.get("items", []):
        name = it.get("name", "")
        qty = it.get("qty", 0)
        rate = it.get("rate", 0)
        amount = it.get("amount", 0)
        sheet.append([bn, dt, name, qty, rate, amount, total, payment, table])
    wb.save(BILLS_XLSX)

def read_bills_from_xlsx():
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX)
    sheet = wb.active
    bills = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        # Bill No, Date & Time, Item Name, Qty, Rate, Amount, Total, Payment Method, Table
        row3 = list(row) + [None]*9
        bill_no, dt, item_name, qty, rate, amount, total, payment, table = row3[:9]
        key = str(bill_no)
        if key not in bills:
            bills[key] = {
                "billNo": bill_no,
                "dateTime": dt,
                "table": table,
                "payment": payment,
                "total": total,
                "items": []
            }
        bills[key]["items"].append({"name": item_name, "qty": qty, "rate": rate, "amount": amount})
    out = list(bills.values())
    try:
        out.sort(key=lambda x: int(x.get("billNo") or 0))
    except:
        pass
    return out

# -----------------------------
# Flask endpoints
# -----------------------------

INDEX_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Tempura — Local Excel Billing</title>
  <style>
    body{font-family:Arial,Helvetica,sans-serif;background:#f4f6f8;margin:0;color:#111}
    header{background:#2c3e50;color:#fff;padding:16px;text-align:center}
    .brand{font-size:26px;font-weight:800;color:#f39c12}
    .container{max-width:1200px;margin:16px auto;padding:0 12px}
    .topbar{display:flex;gap:8px;align-items:center}
    input,select,button{padding:8px;border-radius:6px;border:1px solid #e6eef6}
    button{cursor:pointer;background:#2c3e50;color:#fff;border:0}
    .layout{display:grid;grid-template-columns:240px 1fr 420px;gap:12px;margin-top:14px}
    .card{background:#fff;padding:12px;border-radius:8px;box-shadow:0 8px 20px rgba(0,0,0,0.04)}
    .tableBtn{padding:10px;border:1px solid #eef2f7;border-radius:6px;margin-bottom:8px;cursor:pointer;display:flex;justify-content:space-between}
    .tableBtn.active{background:#fbf6ec;border-color:#f1c40f}
    .items{display:flex;flex-wrap:wrap;gap:8px;margin-top:12px}
    .item{padding:10px;border:1px solid #eef7ff;border-radius:6px;min-width:160px;cursor:pointer;background:#fdfefe}
    table{width:100%;border-collapse:collapse}
    th,td{padding:8px;border-bottom:1px solid #f1f3f6}
    .small{font-size:13px;color:#666}
    .actions{display:flex;gap:8px;margin-top:12px}
    @media(max-width:980px){ .layout{grid-template-columns:1fr} .topbar{flex-direction:column;align-items:flex-start} }
  </style>
</head>
<body>
<header><div class="brand">Tempura</div><div>Local Excel Billing — menu.xlsx & bills.xlsx in project folder</div></header>
<div class="container">
  <div class="topbar">
    <input id="hotelName" placeholder="Hotel name" style="width:260px" value="Tempura Hotel"/>
    <input id="hotelAddr" placeholder="Address" style="width:480px" value="123 Main Road, Pondicherry | Ph: 9876543210"/>
    <div style="flex:1"></div>
    <button onclick="window.location.reload()">Reload</button>
  </div>

  <div class="layout">
    <!-- left: tables -->
    <div class="card">
      <div style="font-weight:800">Tables</div>
      <div class="small">Click to open a table</div>
      <div id="tables" style="margin-top:12px"></div>
      <div class="actions" style="margin-top:12px">
        <button onclick="downloadFile('/download/menu.xlsx')">Download menu.xlsx</button>
        <button onclick="downloadFile('/download/bills.xlsx')">Download bills.xlsx</button>
      </div>
    </div>

    <!-- center: menu items -->
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center">
        <div>
          <div id="activeTableLabel" style="font-weight:800">Select a table</div>
          <div id="hint" class="small">Menu loaded from menu.xlsx (no samples auto-inserted).</div>
        </div>
        <div>
          <select id="categorySelect" onchange="renderItems()"></select>
          <input id="search" placeholder="Search item..." oninput="renderItems()" style="width:220px;margin-left:8px"/>
        </div>
      </div>

      <div id="items" class="items"></div>

      <div style="margin-top:12px">
        <button onclick="toggleAddMenu()">Add Menu</button>
        <span class="small" style="margin-left:8px">Adds directly to menu.xlsx</span>
      </div>

      <div id="addMenu" style="display:none;margin-top:10px">
        <input id="m_cat" placeholder="Category"/>
        <input id="m_item" placeholder="Item name"/>
        <input id="m_price" placeholder="Price" type="number"/>
        <button onclick="addMenu()">Save Menu</button>
      </div>
    </div>

    <!-- right: cart -->
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center">
        <div style="font-weight:800">Cart</div>
        <div id="cartTableLabel" class="small">No table</div>
      </div>

      <div style="margin-top:8px">
        <table id="cartTable"><thead><tr><th>Item</th><th>Qty</th><th>Rate</th><th>Amt</th></tr></thead><tbody></tbody></table>
      </div>

      <div style="display:flex;justify-content:space-between;margin-top:8px">
        <div class="small">Items: <span id="itemsCount">0</span></div>
        <div style="font-weight:800" id="grandTotal">₹0.00</div>
      </div>

      <div class="actions">
        <button onclick="clearCart()">Cancel</button>
        <button onclick="openRaiseModal()">Raise Bill</button>
      </div>

      <div style="margin-top:12px">
        <button onclick="showBills()">See Bills</button>
      </div>
    </div>
  </div>

  <!-- see bills area (same page modal-like) -->
  <div id="billsArea" style="display:none;margin-top:12px" class="card">
    <div style="display:flex;justify-content:space-between;align-items:center">
      <div style="font-weight:800">All Bills</div>
      <div>
        <input id="billFilter" placeholder="Filter (bill/table/item)" oninput="renderBills()" />
        <button onclick="downloadFile('/download/bills.xlsx')">Download XLSX</button>
        <button onclick="document.getElementById('billsArea').style.display='none'">Close</button>
      </div>
    </div>
    <div style="margin-top:8px;overflow:auto">
      <table id="billsTable"><thead><tr><th>Bill No</th><th>DateTime</th><th>Table</th><th>Total</th><th>Payment</th><th>Items</th></tr></thead><tbody></tbody></table>
    </div>
  </div>

</div>

<!-- Raise bill modal (simple in-page section) -->
<div id="raiseModal" style="display:none;position:fixed;left:0;top:0;width:100%;height:100%;background:rgba(0,0,0,0.45);align-items:center;justify-content:center">
  <div style="background:#fff;padding:20px;border-radius:8px;max-width:720px;margin:auto">
    <div style="display:flex;justify-content:space-between;align-items:center">
      <div><strong id="modalTitle">Bill Preview</strong></div>
      <div><button onclick="closeRaiseModal()">Close</button></div>
    </div>
    <div id="printArea" style="margin-top:12px"></div>
    <div style="margin-top:12px;display:flex;gap:8px;justify-content:flex-end">
      <select id="paymentSelect"><option>Cash</option><option>Paytm</option><option>Card</option><option>UPI</option></select>
      <button onclick="printPreview()">Print</button>
      <button onclick="saveBill()">Save Bill</button>
    </div>
  </div>
</div>

<script>
/* Frontend logic */
const TABLES = {{ tables|tojson }};
let MENU = {};
let CARTS = {};
let ACTIVE_TABLE = null;

function init(){
  // make sure server has created files
  fetch('/api/menu').then(r=>r.json()).then(data=>{
    MENU = data || {};
    renderCategories();
    renderItems();
  });

  // setup tables
  const tdiv = document.getElementById('tables');
  tdiv.innerHTML = '';
  TABLES.forEach(t=>{
    CARTS[t] = {};
    const btn = document.createElement('div');
    btn.className = 'tableBtn';
    btn.textContent = t;
    btn.onclick = ()=> selectTable(t, btn);
    tdiv.appendChild(btn);
  });
}

function renderCategories(){
  const sel = document.getElementById('categorySelect');
  sel.innerHTML = '';
  Object.keys(MENU).forEach(cat=>{
    const opt = document.createElement('option'); opt.value = cat; opt.textContent = cat;
    sel.appendChild(opt);
  });
  if(!sel.value) sel.value = Object.keys(MENU)[0] || '';
}

function renderItems(){
  const cat = document.getElementById('categorySelect').value || Object.keys(MENU)[0];
  const q = (document.getElementById('search').value||'').toLowerCase();
  const cont = document.getElementById('items');
  cont.innerHTML = '';
  (MENU[cat]||[]).filter(it => !q || it.name.toLowerCase().includes(q)).forEach(it=>{
    const d = document.createElement('div'); d.className='item';
    d.innerHTML = `<div style="font-weight:700">${it.name}</div><div>₹${Number(it.price).toFixed(2)}</div>`;
    d.onclick = ()=> addItem(it);
    cont.appendChild(d);
  });
}

function selectTable(name, el){
  ACTIVE_TABLE = name;
  document.getElementById('activeTableLabel').textContent = name;
  document.getElementById('cartTableLabel').textContent = name;
  document.querySelectorAll('.tableBtn').forEach(b=> b.classList.toggle('active', b.textContent===name));
  updateCartUI();
}

function addItem(it){
  if(!ACTIVE_TABLE){ alert('Select a table first'); return; }
  const cart = CARTS[ACTIVE_TABLE];
  if(cart[it.name]) cart[it.name].qty += 1;
  else cart[it.name] = { name: it.name, qty: 1, rate: Number(it.price), amount: 0 };
  cart[it.name].amount = cart[it.name].qty * cart[it.name].rate;
  updateCartUI();
}

function updateCartUI(){
  const tbody = document.querySelector('#cartTable tbody');
  tbody.innerHTML = '';
  if(!ACTIVE_TABLE){ tbody.innerHTML = '<tr><td colspan="4">Open a table</td></tr>'; document.getElementById('grandTotal').innerText='₹0.00'; document.getElementById('itemsCount').innerText='0'; return; }
  const cart = CARTS[ACTIVE_TABLE];
  const keys = Object.keys(cart);
  if(keys.length === 0){ tbody.innerHTML = '<tr><td colspan="4">Cart empty</td></tr>'; document.getElementById('grandTotal').innerText='₹0.00'; document.getElementById('itemsCount').innerText='0'; return; }
  let total = 0, count = 0;
  keys.forEach(k=>{
    const it = cart[k];
    const tr = document.createElement('tr');
    tr.innerHTML = `<td>${it.name}</td><td>${it.qty}</td><td>₹${it.rate.toFixed(2)}</td><td>₹${it.amount.toFixed(2)}</td>`;
    tbody.appendChild(tr);
    total += it.amount; count += it.qty;
  });
  document.getElementById('grandTotal').innerText = '₹' + total.toFixed(2);
  document.getElementById('itemsCount').innerText = count;
}

function clearCart(){
  if(!ACTIVE_TABLE) return alert('Open a table');
  if(!confirm('Clear cart for ' + ACTIVE_TABLE + '?')) return;
  CARTS[ACTIVE_TABLE] = {};
  updateCartUI();
}

function toggleAddMenu(){ document.getElementById('addMenu').style.display = (document.getElementById('addMenu').style.display === 'none' ? 'block' : 'none'); }

async function addMenu(){
  const cat = document.getElementById('m_cat').value.trim();
  const name = document.getElementById('m_item').value.trim();
  const price = Number(document.getElementById('m_price').value || 0);
  if(!cat || !name || !price) return alert('Provide category, name and price');
  const payload = { category: cat, name: name, price: price };
  const res = await fetch('/api/menu', { method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(payload)});
  if(res.ok){
    alert('Menu saved to menu.xlsx');
    // reload menu
    const r = await fetch('/api/menu'); MENU = await r.json();
    renderCategories(); renderItems();
    document.getElementById('m_cat').value=''; document.getElementById('m_item').value=''; document.getElementById('m_price').value='';
    document.getElementById('addMenu').style.display='none';
  } else {
    alert('Failed to save menu');
  }
}

function openRaiseModal(){
  if(!ACTIVE_TABLE) return alert('Open a table');
  const cart = CARTS[ACTIVE_TABLE];
  if(Object.keys(cart).length===0) return alert('Cart empty');
  // build HTML
  const printArea = document.getElementById('printArea');
  const hotel = document.getElementById('hotelName').value;
  const addr = document.getElementById('hotelAddr').value;
  let html = `<div style="text-align:center"><h2>${hotel}</h2><div>${addr}</div><hr/></div>`;
  html += `<div><strong>Table: ${ACTIVE_TABLE}</strong></div>`;
  html += `<table style="width:100%;border-collapse:collapse"><thead><tr><th style="text-align:left">S.no</th><th style="text-align:left">Item</th><th style="text-align:right">Qty</th><th style="text-align:right">Rate</th><th style="text-align:right">Amt</th></tr></thead><tbody>`;
  let total = 0; let i = 1;
  Object.values(cart).forEach(it=>{
    html += `<tr><td>${i}</td><td>${it.name}</td><td style="text-align:right">${it.qty}</td><td style="text-align:right">₹${it.rate.toFixed(2)}</td><td style="text-align:right">₹${it.amount.toFixed(2)}</td></tr>`;
    total += it.amount; i++;
  });
  html += `</tbody></table>`;
  html += `<div style="text-align:right;font-weight:800;margin-top:8px">Total: ₹${total.toFixed(2)}</div>`;
  printArea.innerHTML = html;
  document.getElementById('raiseModal').style.display = 'flex';
}

function closeRaiseModal(){ document.getElementById('raiseModal').style.display = 'none'; }

function printPreview(){
  const w = window.open('', '_blank', 'width=600,height=800');
  w.document.write(document.getElementById('printArea').innerHTML);
  w.document.close();
  w.print();
}

async function saveBill(){
  if(!ACTIVE_TABLE) return alert('Open a table');
  const cart = CARTS[ACTIVE_TABLE];
  const items = Object.values(cart).map(it=>({ name: it.name, qty: it.qty, rate: it.rate, amount: it.amount }));
  const total = items.reduce((s,i)=>s+i.amount,0);
  const nextRes = await fetch('/api/next_bill_no'); const j = await nextRes.json(); const billNo = j.next;
  const payload = { billNo: billNo, dateTime: new Date().toISOString(), table: ACTIVE_TABLE, payment: document.getElementById('paymentSelect').value, total: total, items: items };
  const res = await fetch('/api/bills', { method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(payload) });
  if(res.ok){
    alert('Bill saved to bills.xlsx');
    CARTS[ACTIVE_TABLE] = {};
    updateCartUI();
    closeRaiseModal();
  } else {
    alert('Failed to save bill');
  }
}

async function showBills(){
  document.getElementById('billsArea').style.display = 'block';
  await renderBills();
}

async function renderBills(){
  const res = await fetch('/api/bills');
  const all = await res.json();
  const q = (document.getElementById('billFilter').value||'').toLowerCase();
  const tbody = document.querySelector('#billsTable tbody'); tbody.innerHTML = '';
  all.slice().reverse().forEach(b=>{
    const items = (b.items||[]).map(i=> `${i.name}(${i.qty})`).join(', ');
    if(q && !(String(b.billNo).includes(q) || (b.table||'').toLowerCase().includes(q) || items.toLowerCase().includes(q))) return;
    const tr = document.createElement('tr');
    tr.innerHTML = `<td>${b.billNo}</td><td>${b.dateTime}</td><td>${b.table}</td><td>₹${Number(b.total).toFixed(2)}</td><td>${b.payment || ''}</td><td>${items}</td>`;
    tbody.appendChild(tr);
  });
}

function downloadFile(url){
  window.location = url;
}

/* boot */
window.TABLES = {{ tables|tojson }};
init();
</script>
</body>
</html>
"""

# Flask routes
@app.route("/")
def index():
    # embed Python TABLES into template
    return render_template_string(INDEX_HTML, tables=TABLES)

@app.route("/api/menu", methods=["GET", "POST"])
def api_menu():
    if request.method == "GET":
        menu = load_menu_from_xlsx()
        return jsonify(menu)
    else:
        # POST to add a menu item: expects { category, name, price }
        payload = request.get_json(force=True)
        cat = str(payload.get("category", "")).strip()
        name = str(payload.get("name", "")).strip()
        try:
            price = float(payload.get("price", 0))
        except:
            price = 0.0
        if not name:
            return ("Missing name", 400)
        menu = load_menu_from_xlsx()
        menu.setdefault(cat or "Uncategorized", []).append({"name": name, "price": price})
        write_menu_to_xlsx(menu)
        return ("OK", 200)

@app.route("/api/next_bill_no", methods=["GET"])
def api_next():
    n = next_bill_no()
    return jsonify({"next": n})

@app.route("/api/bills", methods=["GET", "POST"])
def api_bills():
    if request.method == "GET":
        bills = read_bills_from_xlsx()
        return jsonify(bills)
    else:
        data = request.get_json(force=True)
        # Validate
        if not isinstance(data.get("items", []), list):
            return ("Invalid items", 400)
        if "billNo" not in data or not data.get("billNo"):
            data["billNo"] = next_bill_no()
        if "dateTime" not in data or not data.get("dateTime"):
            data["dateTime"] = datetime.now().isoformat()
        append_bill_to_xlsx(data)
        return ("OK", 200)

@app.route("/download/<fname>")
def download(fname):
    # allow only menu.xlsx and bills.xlsx
    if fname not in ("menu.xlsx", "bills.xlsx"):
        return ("Forbidden", 403)
    if not Path(fname).exists():
        # create default if missing
        if fname == "menu.xlsx":
            ensure_menu_file()
        else:
            ensure_bills_file()
    return send_file(fname, as_attachment=True)

if __name__ == "__main__":
    # ensure files exist but leave empty (no sample items)
    ensure_menu_file()
    ensure_bills_file()
    # network-local by default
    app.run(host="127.0.0.1", port=5000, debug=True)
